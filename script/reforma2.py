from collections import defaultdict
from datetime import timedelta, datetime, date

from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


points_name_map = {
    #
    'Списание просрочки "VanLav 1" ЗАЛ': 'Аврора',
    'Списания просрочки "VanLav 2"': 'Красная 77',
    'Списания просрочки "VanLav 3"': 'Ставропольская',
    'Списания просрочки "VanLav 4"': 'Панорама',
    #
    'Витрина "Аврора"': 'Аврора',
    'Бар "Красная #77"': 'Красная 77',
    'Бар "Ставропольская 218"': 'Ставропольская',
    'Бар Кругликовская': 'Панорама',
    #
    '1- ЗАЛ Аврора КРД': 'Аврора',
    '2- ЗАЛ 77 Красная КРД': 'Красная 77',
    '3- ЗАЛ Ставропольская КРД': 'Ставропольская',
    '4- ЗАЛ Панорама КРД': 'Панорама',
}


def read_sales(otd):
    workbook = load_workbook('Продажи (2).xlsx', data_only=True)
    sheet = workbook[workbook.sheetnames[0]]

    for row in sheet.iter_rows(min_row=6):
        if row[0].value in points_name_map:
            curr_otd = points_name_map[row[0].value]
        if row[1].value:
            curr_date = row[1].value.date()

        bludo = (row[3].value or '').strip()
        if bludo:
            count = int(row[4].value)
            otd[curr_otd][bludo][curr_date][2] = count



def read_prod(otd):
    workbook = load_workbook('Приготовления ЦЕХ.xlsx', data_only=True)
    sheet = workbook[workbook.sheetnames[0]]

    for row in sheet.iter_rows(min_row=6):
        if row[1].value in points_name_map:
            curr_otd = points_name_map[row[1].value]
        if row[2].value:
            curr_date = row[2].value.date()

        bludo = (row[4].value or '').strip()
        if bludo:
            count = int(row[5].value)
            otd[curr_otd][bludo][curr_date][1] = count

def read_discard(otd):
    workbook = load_workbook('Списания.xlsx', data_only=True)
    sheet = workbook[workbook.sheetnames[0]]

    for row in sheet.iter_rows(min_row=6):
        if row[1].value in points_name_map:
            curr_otd = points_name_map[row[1].value]
        if row[2].value:
            curr_date = row[2].value.date()

        bludo = (row[4].value or '').strip()
        if bludo:
            count = int(row[5].value)
            otd[curr_otd][bludo][curr_date][3] = count


def write(otd: dict[str, dict[str, dict[date, tuple[int, int, int, int]]]]):
    """

    :param otd: otdel -> bludo -> date -> (zayav, production, sale, discard)
    :return:
    """
    workbook = Workbook()

    for otd_name, otd_data in otd.items():
        sheet: Worksheet = workbook.create_sheet(otd_name)

        bludo_by_row = {row: bludo  for row, bludo in enumerate(sorted(otd_data), start=3)}
        min_date = min(min(bludo_data) for bludo_data in otd_data.values())
        max_date = max(max(bludo_data) for bludo_data in otd_data.values())
        all_dates = []
        date_in_period = min_date
        while True:
            all_dates.append(date_in_period)
            date_in_period = date_in_period + timedelta(days=1)
            if date_in_period > max_date:
                break

        iterator = sheet.iter_rows(max_row=max(bludo_by_row), max_col=len(all_dates) * 4 + 3)

        first_row = next(iterator)
        second_row = next(iterator)
        first_row[0].value = 'Блюдо'
        for col, date in enumerate(all_dates, start=0):
            date_col = col * 4 + 1
            first_row[date_col].value = date.strftime('%d-%b')
            for t_offset, t_name in enumerate(['заявлено', 'произведено', 'продано', 'списано']):
                t_col = date_col + t_offset
                second_row[t_col].value = t_name
        # avg_price_col = col + 1
        # first_row[avg_price_col].value = 'Средняя цена'

        for row in iterator:
            bludo = bludo_by_row[row[0].row]
            row[0].value = bludo
            for col, date_to_process in enumerate(all_dates, start=0):
                date_col_offset = 1 + col * 4
                for t_offset, t_val in enumerate(otd_data[bludo][date_to_process]):
                    t_col = date_col_offset + t_offset
                    row[t_col].value = t_val

            # avg_price_col = col + 1
            # row[avg_price_col].value = next(v for v in otd_data[bludo].values())[1]
    workbook.save('Аггрегация.xlsx')


def main():
    # otd -> bludo -> date -> num
    otd: dict[str,dict[str, dict[date, list[int, int, int, int]]]]
    otd = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: ['РЫБА', '', '', ''])))
    read_prod(otd)
    read_sales(otd)
    read_discard(otd)

    write(otd)






if __name__ == '__main__':
    main()