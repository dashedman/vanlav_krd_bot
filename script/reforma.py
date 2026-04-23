from collections import defaultdict
from datetime import timedelta, datetime

from openpyxl.reader.excel import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


def main():
    workbook = load_workbook('Продажи (2).xlsx', data_only=True)
    sheet = workbook[workbook.sheetnames[0]]

    # otd -> bludo -> date -> num
    otd: dict[str,dict[str, dict[datetime, tuple[int, float]]]] = defaultdict(lambda: defaultdict(dict))
    for row in sheet.iter_rows(min_row=6):
        if row[0].value:
            curr_otd = row[0].value
        if row[1].value:
            curr_date = row[1].value.date()

        bludo = row[3].value
        if bludo:
            count = int(row[4].value)
            if count > 0:
                average_price = float(row[5].value)
                otd[curr_otd][bludo][curr_date] = (count, average_price)

    for otd_name, otd_data in otd.items():
        sheet: Worksheet = workbook.create_sheet(otd_name)

        bludo_by_row = {row: bludo  for row, bludo in enumerate(sorted(otd_data), start=2)}
        min_date = min(min(bludo_data) for bludo_data in otd_data.values())
        max_date = max(max(bludo_data) for bludo_data in otd_data.values())
        all_dates = []
        date_in_period = min_date
        while True:
            all_dates.append(date_in_period)
            date_in_period = date_in_period + timedelta(days=1)
            if date_in_period > max_date:
                break

        iterator = sheet.iter_rows(max_row=max(bludo_by_row), max_col=len(all_dates) + 3)

        first_row = next(iterator)
        first_row[0].value = 'Блюдо'
        for col, date in enumerate(all_dates, start=1):
            first_row[col].value = date.strftime('%d-%b')
        avg_price_col = col + 1
        first_row[avg_price_col].value = 'Средняя цена'

        for row in iterator:
            bludo = bludo_by_row[row[0].row]
            row[0].value = bludo
            for col, date_to_process in enumerate(all_dates, start=1):
                cell_data = otd_data.get(bludo, {}).get(date_to_process, None)
                if cell_data:
                    row[col].value = cell_data[0]
            avg_price_col = col + 1
            row[avg_price_col].value = next(v for v in otd_data[bludo].values())[1]
    workbook.save('Продажи_обработано.xlsx')

    print()






if __name__ == '__main__':
    main()